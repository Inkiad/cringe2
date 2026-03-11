#!/usr/bin/env python3
"""
export.py — Export restaurant data to Excel for manual curation.

Usage:
    python scripts/export.py
    python scripts/export.py --out ~/Desktop/restaurants.xlsx
    python scripts/export.py --zip 91344
    python scripts/export.py --status pending

Requires: pip install openpyxl
"""

import argparse
import glob
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data"))

# Column definitions: (field_key, header_label, editable)
COLUMNS = [
    ("id",              "id",              False),
    ("name",            "name",            False),
    ("address",         "address",         False),
    ("google_rating",   "google_rating",   False),
    ("google_count",    "google_count",    False),
    ("status",          "status",          True),
    ("exclude_reason",  "exclude_reason",  True),
    ("neighborhood",    "neighborhood",    True),
    ("cuisine",         "cuisine",         True),
    ("price",           "price",           True),
    ("vibe",            "vibe",            True),
    ("dishes",          "dishes",          True),
    ("personal_rating", "personal_rating", True),
    ("personal_notes",  "personal_notes",  True),
    ("website",         "website",         True),
]

READONLY_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
HEADER_FILL   = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
HEADER_FONT   = Font(color="FFFFFF", bold=True)

COL_WIDTHS = {
    "id": 30, "name": 30, "address": 40, "google_rating": 13,
    "google_count": 13, "status": 10, "exclude_reason": 22,
    "neighborhood": 20, "cuisine": 25, "price": 10, "vibe": 25,
    "dishes": 25, "personal_rating": 15, "personal_notes": 35, "website": 40,
}


def list_to_str(v):
    if isinstance(v, list):
        return ", ".join(v)
    return v if v is not None else ""


def load_zip_data(zip_filter=None, status_filter=None):
    """Load per-zip JSONs. Returns list of (zip_code, [restaurants]) sorted by zip."""
    files = sorted(glob.glob(os.path.join(DATA_DIR, "restaurants-*.json")))
    result = []
    for path in files:
        base = os.path.basename(path)
        if base == "restaurants-all.json":
            continue
        zip_code = base.replace("restaurants-", "").replace(".json", "")
        if zip_filter and zip_code not in zip_filter:
            continue
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        restaurants = data.get("restaurants", [])
        if status_filter:
            restaurants = [r for r in restaurants if r.get("status") == status_filter]
        if restaurants:
            result.append((zip_code, restaurants))
    return result


def write_sheet(wb, zip_code, restaurants):
    ws = wb.create_sheet(title=zip_code)

    # Header row
    for col_idx, (key, label, editable) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(restaurants, 2):
        for col_idx, (key, label, editable) in enumerate(COLUMNS, 1):
            val = list_to_str(r.get(key))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if not editable:
                cell.fill = READONLY_FILL

    # Freeze header row, set column widths
    ws.freeze_panes = "A2"
    for col_idx, (key, label, editable) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(key, 15)


def main():
    parser = argparse.ArgumentParser(description="Export restaurant data to Excel.")
    parser.add_argument("--out", default=None,
                        help="Output path (default: data/restaurants.xlsx)")
    parser.add_argument("--zip", nargs="+", metavar="ZIP",
                        help="Only include these zip codes")
    parser.add_argument("--status", default=None,
                        help="Filter by status: pending / kept / excluded")
    args = parser.parse_args()

    out_path = args.out or os.path.join(DATA_DIR, "restaurants.xlsx")
    zip_filter = set(args.zip) if args.zip else None

    zip_data = load_zip_data(zip_filter, args.status)
    if not zip_data:
        print("No data found. Run fetch.py first.")
        raise SystemExit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    total = 0
    for zip_code, restaurants in zip_data:
        write_sheet(wb, zip_code, restaurants)
        total += len(restaurants)
        print(f"  {zip_code}: {len(restaurants)} restaurants")

    wb.save(out_path)
    print(f"\nSaved {total} restaurants across {len(zip_data)} zips to {out_path}")


if __name__ == "__main__":
    main()
