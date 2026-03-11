#!/usr/bin/env python3
"""
import.py — Import Excel curation back into per-zip JSONs and generate restaurants-all.json.

Usage:
    python scripts/import.py
    python scripts/import.py --file data/restaurants.xlsx

Requires: pip install openpyxl
"""

import argparse
import datetime
import json
import os

import openpyxl

DATA_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data"))

EDITABLE_FIELDS = {
    "status", "exclude_reason", "neighborhood",
    "cuisine", "price", "vibe", "dishes",
    "personal_rating", "personal_notes", "website",
}
LIST_FIELDS = {"cuisine", "vibe", "dishes"}


def parse_list(val):
    """Parse comma-separated string to list."""
    if not val:
        return []
    return [x.strip() for x in str(val).split(",") if x.strip()]


def parse_value(key, val):
    if val is None or val == "":
        return [] if key in LIST_FIELDS else None
    if key in LIST_FIELDS:
        return parse_list(val)
    if key == "personal_rating":
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    return val


def load_zip_json(zip_code):
    path = os.path.join(DATA_DIR, f"restaurants-{zip_code}.json")
    if not os.path.exists(path):
        return None, path
    with open(path, encoding="utf-8") as f:
        return json.load(f), path


def save_zip_json(path, meta, restaurants):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "restaurants": restaurants}, f, indent=2, ensure_ascii=False)


def process_sheet(ws):
    """Parse worksheet rows. Returns (zip_code, list of row dicts)."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {h: row[i] if i < len(row) else None
                    for i, h in enumerate(headers) if h}
        if row_dict.get("id"):
            rows.append(row_dict)
    return ws.title, rows


def main():
    parser = argparse.ArgumentParser(
        description="Import Excel curation into per-zip JSONs and generate restaurants-all.json."
    )
    parser.add_argument("--file", default=os.path.join(DATA_DIR, "restaurants.xlsx"),
                        help="Input Excel file (default: data/restaurants.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"Error: file not found: {args.file}")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(args.file)
    all_kept = []
    total_updated = total_unmatched = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        zip_code, rows = process_sheet(ws)

        existing_data, path = load_zip_json(zip_code)
        if existing_data is None:
            print(f"  {zip_code}: no JSON file found — skipping")
            continue

        meta = existing_data.get("meta", {"zip": zip_code})
        meta["last_imported"] = datetime.date.today().isoformat()
        by_id = {r["id"]: r for r in existing_data.get("restaurants", [])}

        updated = unmatched = 0
        for row in rows:
            pid = row.get("id")
            if pid not in by_id:
                unmatched += 1
                continue
            entry = by_id[pid]
            for field in EDITABLE_FIELDS:
                if field in row:
                    entry[field] = parse_value(field, row[field])
            updated += 1

        restaurants = sorted(by_id.values(), key=lambda r: r["name"].lower())
        save_zip_json(path, meta, restaurants)

        kept_here = [r for r in restaurants if r.get("status") == "kept"]
        all_kept.extend(kept_here)
        print(f"  {zip_code}: {updated} updated, {unmatched} unmatched, {len(kept_here)} kept")
        total_updated += updated
        total_unmatched += unmatched

    # Generate restaurants-all.json (kept only, sorted by google_rating desc)
    all_kept.sort(key=lambda r: -(r.get("google_rating") or 0))
    all_path = os.path.join(DATA_DIR, "restaurants-all.json")
    with open(all_path, "w", encoding="utf-8") as f:
        json.dump(
            {"meta": {"generated": datetime.date.today().isoformat(), "total": len(all_kept)},
             "restaurants": all_kept},
            f, indent=2, ensure_ascii=False,
        )

    print(f"\nTotal: {total_updated} updated, {total_unmatched} unmatched")
    print(f"restaurants-all.json: {len(all_kept)} kept entries")


if __name__ == "__main__":
    main()
